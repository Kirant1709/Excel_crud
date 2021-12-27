import openpyxl
import pandas as pd
import os

#Loading the excel sheet
def load_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    else:
        return "File not found ..."
    
wb_path = "Employee_details.xlsx"
wb = load_workbook(wb_path)
sheet = wb["Sheet1"]
sheet_obj = wb.active
max_column = sheet_obj.max_column
max_row = sheet_obj.max_row


# function to View all Employees
def view_all_employee():
    employee_list = pd.read_excel(wb_path)
    print(employee_list)


#function to Create a new employee
def create_Employee():
    new_employee = input("\n Enter the Employee ID, Name, Year of Joining:- ").split(' ')
    sheet.append(new_employee)
    wb.save(wb_path)
    print("Employee details added successfully")
    add_more = input("\n Add more employees? Yes/No:- ")
    if add_more.lower() == "yes":
        create_Employee()
        

#function to Search employee
def search(name):
    for i in range(1,max_row+1):
        if sheet.cell(row=i,column=2).value == name:
            print("Employee found")
            return i

#function to display employee
def display_employee(row):
    for i in range(1,max_column+1):
        cell_obj = sheet_obj.cell(row = row , column = i)
        print(cell_obj.value)

#function to update employee 
def update_employee(row):
    x = input("\n Enter the Employee ID, Name, Year of Joining:-  ").split(' ')
    for col_index,value in enumerate(x,start=1):
        sheet.cell(row = row, column =col_index, value =value)
        wb.save(wb_path)
    print("Employee Details Updated successfully")   

#function to delete employee
def delete_employee(row):
    sheet.delete_rows(row)
    wb.save(wb_path)
    print("Employee deleted successfully")
        

while True:
    print("\n Welcome to Employee crud operation")
    print("\n 1.View all employees ")
    print("\n 2.Create Employee")
    print("\n 3.Update Employee")
    print("\n 4.Delete Employee")
    ch = input("\n Enter the option:- ")
    if ch == '1':
        view_all_employee()
    if ch == '2':
        create_Employee()        
    if ch == '3':
        x = input("\n Enter the employee name :- ")
        row = search(x)
        display_employee(row)
        y = input("\n Edit employee details ? yes/no:- ")
        if y == 'yes':
            update_employee(row)
    if ch == '4':
        x = input("\n Enter the employee name :")
        row = search(x)
        display_employee(row)
        y = input("\n Delete selected employee ? yes/no:- ")
        if y == 'yes':
            delete_employee(row)
    else:
        break


